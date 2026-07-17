import os
from pathlib import Path
from openpyxl import load_workbook
import urllib.request

# 取得專案根目錄 (OH_H)
BASE_DIR = Path(__file__).resolve().parent.parent
# 定位根目錄下的 data.xlsx 絕對路徑
EXCEL_PATH = os.path.join(BASE_DIR, "data.xlsx")

def load():
    try:
        url = "https://docs.google.com/spreadsheets/d/1svRcTy-waZP1bpnVj0K3jTsIrUHeDLVMHR1Wk47z8To/export?format=xlsx&gid=0"
        # 修正：確保下載到根目錄
        urllib.request.urlretrieve(url, EXCEL_PATH)
    except Exception as e:
        # debug 用的話可以印出來，上線時再 pass
        print(f"雲端更新失敗，使用本地舊快取: {e}")
        pass

# 每次呼叫這個模組時自動跑更新
load()

# 修正：確保讀取的是根目錄下的正確路徑
wb = load_workbook(EXCEL_PATH, read_only=True)
sheet = wb.active

chemicals = {}
for row in sheet.iter_rows(min_row=2, values_only=True):
    # 防止 Excel 下方有空白行導致解包錯誤 (Unpack error)
    if not row[0]: 
        continue
        
    Name, ion, H, OH = row
    chemicals[Name] = {"ion": ion, "H+": H, "OH-": OH, "H+_need": OH}